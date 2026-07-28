#!/usr/bin/env python3
"""
LC Core — client workbook (.xlsx) -> importer-ready CSVs, config-driven.

Generalized from the predecessor project's companion script.
All site knowledge lives in a per-site YAML/JSON config (see
example-normalize-config.yml): which sheets to read, how source columns map to
output columns, taxonomy alias tables, and which columns must be URLs.

Guarantees carried over from production:
  * MANDATORY header-title assertion — if any declared source column title is
    missing from a sheet, the run aborts loudly (exit != 0). Column titles,
    never positions: positional mapping corrupted two imports when the client
    inserted columns.
  * Identity = TITLE ONLY (import_key = md5(title)[:12]) — never a correctable
    field like URL (the 2026-07-09 26-duplicate incident).
  * Post-run URL sanity checks (url_columns) and title-collision warnings.

Usage:
    python3 normalize_workbook.py <workbook.xlsx> <config.yml|config.json> [--out DIR]

Re-runnable; deterministic output (rows sorted by title). Writes one CSV per
configured sheet-output plus normalization-warnings.txt in the output dir.
"""

import argparse
import csv
import json
import itertools
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from normalize_lib import (  # noqa: E402
    DEFAULT_SPLIT_RE,
    apply_row_rules,
    build_lookup,
    find_header_row,
    map_row,
    norm,
    spec_required_titles,
    spreadsheet_safe,
    upsert_record,
)


def load_config(path):
    """Load YAML or JSON config. YAML needs pyyaml; JSON always works."""
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    if path.lower().endswith(".json"):
        return json.loads(text)
    try:
        import yaml
    except ImportError:
        sys.exit(
            "FATAL: config '%s' looks like YAML but pyyaml is not installed.\n"
            "Install it (pip3 install pyyaml) or provide the config as .json." % path
        )
    return yaml.safe_load(text)


def validate_config(cfg):
    """Fail loudly on a malformed config before touching the workbook."""
    problems = []
    if not isinstance(cfg, dict):
        problems.append("config root must be a mapping")
        return problems
    sheets = cfg.get("sheets")
    if not sheets or not isinstance(sheets, list):
        problems.append("config needs a non-empty 'sheets' list")
        return problems
    for i, s in enumerate(sheets):
        where = "sheets[%d]" % i
        if not isinstance(s, dict):
            problems.append(where + " must be a mapping")
            continue
        if not s.get("sheet"):
            problems.append(where + " is missing 'sheet' (source sheet name)")
        if not s.get("output"):
            problems.append(where + " is missing 'output' (CSV filename)")
        cols = s.get("columns")
        if not cols or not isinstance(cols, dict):
            problems.append(where + " needs a 'columns' mapping")
            continue
        if "post_title" not in cols:
            problems.append(where + ".columns must include 'post_title' (identity depends on it)")
        for out_col, spec in cols.items():
            spec = spec or {}
            if not isinstance(spec, dict):
                problems.append("%s.columns.%s must be a mapping" % (where, out_col))
                continue
            keys = {"from", "const", "derive", "tax", "bool"}
            if not (set(spec) & keys):
                problems.append(
                    "%s.columns.%s needs one of: from / const / derive" % (where, out_col)
                )
            if spec.get("derive") not in (None, "title"):
                problems.append("%s.columns.%s: only derive: title is supported" % (where, out_col))
            if spec.get("tax") and "from" not in spec:
                problems.append("%s.columns.%s: 'tax' requires 'from'" % (where, out_col))
        for col in (s.get("url_columns") or []) + (s.get("draft_when_empty") or []):
            if col not in cols:
                problems.append(
                    "%s: '%s' referenced in url_columns/draft_when_empty but not in columns" % (where, col)
                )
    return problems


def rows_of(ws):
    """Yield non-empty rows as lists of normalized strings."""
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            yield [norm(c) for c in row]


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", help="client .xlsx workbook")
    ap.add_argument("config", help="per-site normalize config (.yml/.yaml/.json)")
    ap.add_argument("--out", default=None, help="output dir (default: ./data next to this script)")
    ap.add_argument(
        "--spreadsheet-safe",
        action="store_true",
        help="prefix formula-like CSV cells for safe human opening (do not use for direct WordPress import)",
    )
    args = ap.parse_args(argv)

    cfg = load_config(args.config)
    problems = validate_config(cfg)
    if problems:
        sys.exit("FATAL: invalid normalize config:\n  - " + "\n  - ".join(problems))

    out_dir = args.out or os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(out_dir, exist_ok=True)

    try:
        import openpyxl
    except ImportError:
        sys.exit("FATAL: openpyxl is required to read .xlsx workbooks (pip3 install openpyxl).")

    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    lookup = build_lookup(cfg.get("taxonomy_aliases", {}))
    split_re = cfg.get("split_regex", DEFAULT_SPLIT_RE)
    all_warnings = []

    for sheet_cfg in cfg["sheets"]:
        sheet_name = sheet_cfg["sheet"]
        if sheet_name not in wb.sheetnames:
            sys.exit(
                "FATAL: sheet '%s' not found in workbook. Sheets present: %s"
                % (sheet_name, ", ".join(wb.sheetnames))
            )
        ws = wb[sheet_name]
        columns = sheet_cfg["columns"]
        required = spec_required_titles(columns)

        rows = rows_of(ws)
        header_scan = list(itertools.islice(rows, 20))
        # MANDATORY header-title assertion — HeaderError is a SystemExit (loud, non-zero).
        header_idx, header_map = find_header_row(header_scan, required)
        print("[%s] header validated (%d source columns) at row %d" % (sheet_name, len(required), header_idx + 1))

        records = {}
        data_rows = itertools.chain(header_scan[header_idx + 1 :], rows)
        for row_cells in data_rows:
            record, warnings = map_row(row_cells, header_map, columns, lookup, split_re)
            record = apply_row_rules(record, sheet_cfg, warnings)
            for w in warnings:
                all_warnings.append("%s: %s" % (sheet_name, w))
            upsert_record(records, record, all_warnings)

        out_path = os.path.join(out_dir, sheet_cfg["output"])
        fieldnames = list(columns.keys())
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            w.writeheader()
            for record in sorted(records.values(), key=lambda r: r.get("post_title", "").lower()):
                output_record = (
                    {key: spreadsheet_safe(value) for key, value in record.items()}
                    if args.spreadsheet_safe
                    else record
                )
                w.writerow(output_record)
        print("[%s] wrote %d rows -> %s" % (sheet_name, len(records), out_path))

    all_warnings = list(dict.fromkeys(all_warnings))  # dedupe, keep order
    warn_path = os.path.join(out_dir, "normalization-warnings.txt")
    with open(warn_path, "w", encoding="utf-8") as f:
        f.write("\n".join(all_warnings) or "none")
    print("warnings: %d (see %s)" % (len(all_warnings), warn_path))


if __name__ == "__main__":
    main()
